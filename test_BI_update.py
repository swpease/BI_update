import unittest
import os

import openpyxl

import BI_update


class TestGetData(unittest.TestCase):
    def test_wrong_filename(self):
        self.assertRaises(FileNotFoundError, BI_update.get_data, 'wrongfile')

    def test_import(self):
        test_data = [
            ['Item Sales', 'Bradshow', 'Order Start Date:', ' 01/01/16', 'Order End Date:', ' 01/02/16',
             'Item (begins with):\n', 'All', 'Brand  (begins with):\n', 'All', 'Category  (begins with):\n', 'All',
             'Last Name', 'First Name', 'Order Number', 'Invoice Date', 'Quantity'],
            ['11ROZV', '2011  Roussanne', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
            ['Customer', 'Retail', 75632.0, ' 01/02/16   PM', 1.0, '', '', '', '', '', '', '', '', '', '', '', ''],
            ['Customer', 'Retail', 75642.0, ' 01/02/16   PM', 12.0, '', '', '', '', '', '', '', '', '', '', '', ''],
            ['Smith', 'Dave', 75650.0, ' 01/02/16   PM', 6.0, '', '', '', '', '', '', '', '', '', '', '', ''],
            ['11ROZV', 'Count of Customers', 5.0, 26.0, '', '', '', '', '', '', '', '', '', '', '', '', ''],
            ['11SYMESA', '2011 Mesa Reserve Syrah', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
            ['Brown', 'John', 75585.0, ' 01/02/16   PM', 1.0, '', '', '', '', '', '', '', '', '', '', '', ''],
            ['Customer', 'Retail', 75676.0, ' 01/02/16   PM', 1.0, '', '', '', '', '', '', '', '', '', '', '', ''],
            ['SAMPLES', 'Pours', 75678.0, ' 01/02/16   PM', 3.0, '', '', '', '', '', '', '', '', '', '', '', ''],
            ['SAMPLES', '500 - TR', 75683.0, ' 01/02/16   PM', 2.0, '', '', '', '', '', '', '', '', '', '', '', ''],
            ['11SYMESA', 'Count of Customers', 6.0, 11.0, '', '', '', '', '', '', '', '', '', '', '', '', ''],
            ['Page -1 of 1', 'Scott Pease', 'Printed 2/3/2017 11:26:49AM', '', '', '', '', '', '', '', '', '', '', '',
             '', '', '']]

        self.assertEqual(BI_update.get_data('/BI_update/ExcelFiles/itemsales_FORTESTING.xls'), test_data)

    def test_multisheet_import(self):
        # Different because the second sheet doesn't have as many cols (xlrd makes an mxn matrix)
        test_data = [
            ['Item Sales', 'Bradshow', 'Order Start Date:', ' 01/01/16', 'Order End Date:', ' 01/02/16',
             'Item (begins with):\n', 'All', 'Brand  (begins with):\n', 'All', 'Category  (begins with):\n', 'All',
             'Last Name', 'First Name', 'Order Number', 'Invoice Date', 'Quantity'],
            ['11ROZV', '2011  Roussanne', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
            ['Customer', 'Retail', 75632.0, ' 01/02/16   PM', 1.0, '', '', '', '', '', '', '', '', '', '', '', ''],
            ['Customer', 'Retail', 75642.0, ' 01/02/16   PM', 12.0, '', '', '', '', '', '', '', '', '', '', '', ''],
            ['Smith', 'Dave', 75650.0, ' 01/02/16   PM', 6.0, '', '', '', '', '', '', '', '', '', '', '', ''],
            ['11ROZV', 'Count of Customers', 5.0, 26.0, '', '', '', '', '', '', '', '', '', '', '', '', ''],
            ['11SYMESA', '2011 Mesa Reserve Syrah', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''],
            ['Brown', 'John', 75585.0, ' 01/02/16   PM', 1.0],
            ['Customer', 'Retail', 75676.0, ' 01/02/16   PM', 1.0],
            ['SAMPLES', 'Pours', 75678.0, ' 01/02/16   PM', 3.0],
            ['SAMPLES', '500 - TR', 75683.0, ' 01/02/16   PM', 2.0],
            ['11SYMESA', 'Count of Customers', 6.0, 11.0, ''],
            ['Page -1 of 1', 'Scott Pease', 'Printed 2/3/2017 11:26:49AM', '', '']]

        self.assertEqual(BI_update.get_data('/BI_update/ExcelFiles/itemsales_multipage_FORTESTING.xls'), test_data)


class TestSave(unittest.TestCase):

    @classmethod
    def tearDownClass(cls):
        os.remove('../TESTsave.xlsx')

    def test_save(self):
        BI_update.save([[1,2],[3,4]], 'TESTsave')
        saved_data = openpyxl.load_workbook('../TESTsave.xlsx')
        saved_sheet = saved_data.active

        manual_save_comp = openpyxl.load_workbook('ExcelFiles/manual_save_FORTESTING.xlsx')
        manual_sheet = manual_save_comp.active

        for row in manual_sheet.rows:
            for cell in row:
                self.assertEqual(cell.value, saved_sheet[cell.coordinate].value)


class TestProcessItemSales(unittest.TestCase):

    @classmethod
    def tearDownClass(cls):
        os.remove('../pythonISOutputTEST.xlsx')

    def test_processes_correctly(self):
        BI_update.process_item_sales('BI_update/ExcelFiles/itemsales_FORTESTING.xls', 'pythonISOutputTEST')
        auto_generated_wb = openpyxl.load_workbook('../pythonISOutputTEST.xlsx')
        auto_generated_sheet = auto_generated_wb.active

        manually_made_wb = openpyxl.load_workbook(filename='ExcelFiles/manually_processed_itemsalesFORTESTING.xlsx')
        manually_made_sheet = manually_made_wb['Sheet1']

        for row in manually_made_sheet.rows:
            for cell in row:
                self.assertEqual(cell.value, auto_generated_sheet[cell.coordinate].value)

    def test_bad_sku(self):
        self.assertRaises(KeyError, BI_update.process_item_sales,
                          'BI_update/ExcelFiles/itemsales_badsku_FORTESTING.xls',
                          'pythonISOutputTEST')


class TestProcessInvoiceDetails(unittest.TestCase):

    @classmethod
    def tearDownClass(cls):
        os.remove('../pythonIDOutputTEST.xlsx')

    def test_processes_correctly(self):
        BI_update.process_invoice_details('BI_update/ExcelFiles/directmarketingreporttransactionitem_FORTESTING.xls',
                                          'pythonIDOutputTEST')
        auto_generated_wb = openpyxl.load_workbook('../pythonIDOutputTEST.xlsx')
        auto_generated_sheet = auto_generated_wb.active

        manually_made_wb = openpyxl.load_workbook(filename='ExcelFiles/manually_processed_directmarketingreporttransactionitem_FORTESTING.xlsx')
        manually_made_sheet = manually_made_wb['Sheet1']

        for row in manually_made_sheet.rows:
            for cell in row:
                self.assertEqual(cell.value, auto_generated_sheet[cell.coordinate].value)

    def test_bad_sku(self):
        self.assertRaises(KeyError, BI_update.process_invoice_details,
                          'BI_update/ExcelFiles/directmarketingreporttransactionitem_badsku_FORTESTING.xls',
                          'pythonIDOutputTEST')


class TestMain(unittest.TestCase):
    
    def test_main_cleanup(self):
        sales_input_filename = 'BI_update/ExcelFiles/itemsales_FORTESTING.xls'
        invoice_input_filename = 'BI_update/ExcelFiles/directmarketingreporttransactionitem_FORTESTING.xls'
        item_sales_output_filename = 'ISOUT'
        invoice_details_output_filename = 'IDOUT'
        
        self.assertRaises(FileNotFoundError, openpyxl.load_workbook,
                          '../ISOUT')
        self.assertRaises(FileNotFoundError, openpyxl.load_workbook,
                          '../IDOUT')


if __name__ == '__main__':
    unittest.main()